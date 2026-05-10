import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "用户导入模板"

# 表头
headers = ["用户名*", "邮箱*", "密码*", "昵称", "角色(0/1)"]
sheet.append(headers)

# 表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_num, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 示例数据
examples = [
    ["zhangsan", "zhangsan@example.com", "123456", "张三", "0"],
    ["lisi", "lisi@example.com", "123456", "李四", "0"],
    ["admin01", "admin01@example.com", "123456", "管理员", "1"]
]
for example in examples:
    sheet.append(example)
    for col_num in range(1, len(headers) + 1):
        cell = sheet.cell(row=sheet.max_row, column=col_num)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 调整列宽
sheet.column_dimensions['A'].width = 18
sheet.column_dimensions['B'].width = 28
sheet.column_dimensions['C'].width = 18
sheet.column_dimensions['D'].width = 18
sheet.column_dimensions['E'].width = 14

# 添加说明sheet
guide_sheet = workbook.create_sheet("填写说明")
guide_sheet.append(["字段", "说明", "必填"])
guide_sheet.append(["用户名", "3-50位，只能包含字母、数字和下划线", "是"])
guide_sheet.append(["邮箱", "有效的邮箱地址", "是"])
guide_sheet.append(["密码", "6-32位字符", "是"])
guide_sheet.append(["昵称", "用户显示名称，不填则默认使用用户名", "否"])
guide_sheet.append(["角色", "0=普通用户，1=管理员，不填默认为0", "否"])
guide_sheet.append([])
guide_sheet.append(["注意：带 * 号的列为必填项"])
guide_sheet.append(["第一行为表头，数据从第二行开始填写"])

for col in ['A', 'B', 'C']:
    guide_sheet.column_dimensions[col].width = 35

workbook.save("templates/用户导入模板.xlsx")
print("模板文件已生成: templates/用户导入模板.xlsx")
